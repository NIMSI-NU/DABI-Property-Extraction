import jax
import jax.numpy as np
import os
import numpy as onp
from numpy import outer
from openpyxl import Workbook, load_workbook
from Solver.problem_new import Problem
import jax.flatten_util
import glob

from jax import config, jit
import logging
import jax.lax.linalg as lax_linalg
from jax import custom_jvp
from functools import partial

from jax import lax
from jax.numpy.linalg import solve
import time
from jax_fem.utils import save_sol
from implementations.abaqus_input_funcs import *


EPSILON = 1e-12

class OptimizationLogger:
    """
    Lightweight optimization logger.

    Changes vs. original:
    - Epoch rows are accumulated in memory and flushed to the Excel workbook
      only when flush_to_excel() is called (i.e. at the end of each outer loop
      or on explicit request).  The original code called load_workbook() +
      wb.save() on every epoch, which re-parses and re-serialises the entire
      XML for every function evaluation — measurable overhead for 30+ epochs.
    - A plain-text CSV is written incrementally so progress is always visible
      on disk even before the Excel flush.
    """
    def __init__(self, output_dir, num_z_params=6, outer_loop=0):
        self.output_dir = os.path.join(output_dir, "History_%i" % outer_loop)
        os.makedirs(self.output_dir, exist_ok=True)
        self.num_z_params = num_z_params
        self.log_filepath    = os.path.join(self.output_dir, "History_%i_optimization_log.xlsx" % outer_loop)
        self.log_currentJ    = os.path.join(self.output_dir, "History_%i_J.txt" % outer_loop)
        self.log_csv         = os.path.join(self.output_dir, "History_%i_optimization_log.csv" % outer_loop)
        self._pending_rows: list = []
        self._init_logs()

    def _init_logs(self):
        # Initialise a fresh Excel workbook (headers only)
        wb = Workbook()
        ws = wb.active
        headers = ["Epoch", "Objective", "Duration"] + [f"Rho_{i}" for i in range(self.num_z_params)]
        ws.append(headers)
        wb.save(self.log_filepath)
        # Initialise CSV with the same headers
        with open(self.log_csv, 'w', newline='') as f:
            f.write(",".join(headers) + "\n")

    def log_iteration(self, epoch, J, duration, z):
        # Update the current-J sentinel (cheap plain-text write)
        with open(self.log_currentJ, mode='w', newline='') as f:
            f.write(f"{J:.6e}\n")

        row_data = [epoch, float(J), float(duration)] + [float(z[i]) for i in range(self.num_z_params)]

        # Accumulate for deferred Excel flush
        self._pending_rows.append(row_data)

        # Always write to CSV immediately so results are on disk even mid-run
        with open(self.log_csv, mode='a', newline='') as f:
            f.write(",".join(str(v) for v in row_data) + "\n")

    def flush_to_excel(self):
        """Write all buffered rows to the Excel workbook.  Call at end of outer loop."""
        if not self._pending_rows:
            return
        wb = load_workbook(self.log_filepath)
        ws = wb.active
        for row in self._pending_rows:
            ws.append(row)
        wb.save(self.log_filepath)
        self._pending_rows.clear()

    def read_current_objective(self):
        if os.path.exists(self.log_currentJ):
            with open(self.log_currentJ, mode='r') as f:
                return float(f.read().strip())
        return None

class OptimizationProblem:
    def __init__(self, J_total_func, logger_obj, data_dir, data_name, problem, logger, outer_loop=0, max_iters=60, min_cost=1e-2):
        self.J_total_func = J_total_func
        self.logger = logger_obj
        self.counter = 0
        self.dJ = None
        self.data_dir = data_dir
        self.data_name = data_name
        self.problem = problem
        self.outer_loop = outer_loop
        self.info_logger = logger
        self.max_epochs = max_iters
        self.min_cost = min_cost
        self.best_info = None
        self.best_cost = None
        self.best_rho = None
        logger.info(f"#################   INC LOOP: {outer_loop}   ################# \n")
        problem.best_file.write(f"#################   INC LOOP: {outer_loop}   ################# \n")
        
    def objective(self, z):
        rho_ini = z
        logger = self.info_logger
        s = f"\n================ EPOCH {self.counter + 1} ================\n"
        start_time = time.time()
        J, dJ_z, aux = self.J_total_func(z)
        self.dJ = onp.array(dJ_z, dtype=onp.float64)
        problem = self.problem
        duration = time.time() - start_time
        self.logger.log_iteration(self.counter + 1, J, duration, z)
        
        # --- CHECK IF THIS IS THE BEST RESULT SO FAR ---
        is_best = (not hasattr(self, 'best_cost')) or (self.best_cost is None) or (J < self.best_cost)

        # --- FORMAT AND PRINT THE DIAGNOSTICS ---
        if 'sol' in aux and len(aux['sol']) > 0:
            current_dir = os.path.join(self.data_dir, 'History_%i' % (self.outer_loop), 'Epoch_%i' % (self.counter + 1))
            os.makedirs(current_dir, exist_ok=True)
            J_list_array = onp.array(aux['J_list'])
            rho_ini = aux['rho_ini']
            s += f"Rho_ini: {rho_ini}\n"
            
            with open("%s/metrics.csv" % current_dir, 'w') as f:
                f.write("Step,Time_s,Pressure,J_Total,Max_Target_Disp,Max_Sol_Disp,Max_Plastic_Strain,Max_Creep_Strain,Max_VM\n")
                node_inds = aux['target_nodes']
                
                # We still loop to write the CSV and optionally save individual VTUs
                for i in range(len(aux['sol'])):
                    sol = aux['sol'][i]
                    vm = aux['vm'][i]
                    pressure_val = float(aux['pressures'][i]) * np.abs(problem.pressure_mag)
                    save_pressure = pressure_val * np.ones_like(vm)
                    time_val = float(aux['times'][i])
                    j_total_val = float(J_list_array[i])
                    max_tar_disp = float(np.max(aux['tar'][i]))
                    max_sol_disp = float(np.max(sol[node_inds, -1]))
                    max_p_strain = float(np.max(aux['p_strain'][i]))
                    max_c_strain = float(np.max(aux['c_strain'][i]))
                    max_vm_val = float(np.max(aux['vm'][i]))

                    s += f"Step = {i + 1}, Pressure = {pressure_val:.4f}, Target = {max_tar_disp:.6e}, Sol = "
                    s += f"{max_sol_disp:.6e}, Plastic strain = {max_p_strain:.6e}, Creep strain = {max_c_strain:.6e}\n"
                    row = f"{i + 1},{time_val:.4f},{pressure_val:.4f},{j_total_val:.10e},{max_tar_disp:.6e},{max_sol_disp:.6e},{max_p_strain:.6e},{max_c_strain:.6e},{max_vm_val:.4f}\n"
                    f.write(row)
                    
                    # IF BEST, generate the VTUs individually for ParaView
                    if is_best:
                        vtk_path = os.path.join(current_dir, f'u_{i:03d}.vtu')
                        save_sol(self.problem.fe, np.concatenate([sol, np.zeros((sol.shape[0], 1), dtype=sol.dtype)], axis=1), vtk_path, 
                                cell_infos=[('Stress: von Mises', vm), ('Pressure', save_pressure), ('Plastic Strain', aux['p_strain'][i]), 
                                            ('Creep Strain', aux['c_strain'][i]), ('Stress', full_ordered_np_arr(aux['stress'][i])), 
                                            ('Logarithmic Strain', full_ordered_np_arr(aux['strain'][i]))])

            # === SAVE EVERYTHING IN ONE COMPRESSED NPZ OUTSIDE THE LOOP ===
            # This replaces writing dozens of separate .npz files with just one.
            npz_path = os.path.join(current_dir, 'epoch_data.npz')
            
            # Use list comprehensions for custom array functions to process all timesteps cleanly
            onp.savez_compressed(npz_path, sol=aux['sol'], vm=aux['vm'], pressures=aux['pressures'],
                                 p_strain=aux['p_strain'], c_strain=aux['c_strain'],
                                 stress=np.array([full_ordered_np_arr(p) for p in aux['stress']]),
                                 strain=np.array([full_ordered_np_arr(e) for e in aux['strain']]))

            s_header = ['E', 'YS', 'Q', 'b', 'A', 'n']
            rho_ini_mod = (rho_ini + 1) / 2
            for i in range(len(s_header)):
                if i == 0:
                    s += f"Rho {i}: {rho_ini[i]:.6f}; {s_header[i]}; {(problem.param[i][0] + problem.param[i][1] * rho_ini_mod[i]) / 1e3}\n"
                elif i == 4:
                    s += f"Rho {i}: {rho_ini[i]:.6f}; {s_header[i]}; {10 ** (problem.param[4][0] + rho_ini_mod[i] * problem.param[4][1])}\n"
                else:
                    s += f"Rho {i}: {rho_ini[i]:.6f}; {s_header[i]}; {problem.param[i][0] + problem.param[i][1] * rho_ini_mod[i]}\n"
            s += f"J Total; {float(J):.10f}\n"
            s += f"Max Plastic Strain; {float(max_p_strain):.10f}\n"
            s += f"Max Creep Strain; {float(max_c_strain):.10f}\n"
            s += f"J list: {J_list_array}\n"
            s += f"dJ Total: {self.dJ}\n"
        else:
            s += f"J Total: {J:.10f}\nSolver likely diverged or failed to reach time markers.\n"
            current_dir = None
        
        logger.info(s)
        
        # --- CLEANUP OLD BEST & UPDATE METRICS ---
        if is_best:
            if hasattr(self, 'best_dir') and self.best_dir and os.path.exists(self.best_dir):
                # Find all .vtu files in the old best directory and delete them
                old_vtus = glob.glob(os.path.join(self.best_dir, '*.vtu'))
                for vtk_f in old_vtus:
                    try:
                        os.remove(vtk_f)
                    except OSError:
                        pass # Handle file-locked or missing edge cases gracefully
            
            # 2. Record the new global milestones
            self.best_cost = J
            self.best_info = s
            self.best_rho = rho_ini
            self.best_dir = current_dir # Track where the current VTUs live
            
        self.counter += 1
        if self.counter % 5 == 0:
            self.logger.flush_to_excel()
        return onp.array(J, dtype=onp.float64)
    
    def gradient(self, z):
        return self.dJ
    
    def early_stopping_callback(self, xk, *args, **kwargs):
        if self.counter >= self.max_epochs:
            self.info_logger.info(f"Early stopping triggered: {self.max_epochs} epochs reached.")
            raise StopIteration
        if self.counter > self.max_epochs / 2 and self.best_cost is not None and self.best_cost < self.min_cost:
            self.info_logger.info(f"Early stopping triggered: {self.min_cost} cost reached.")
            raise StopIteration

def safe_divide(x, y):
    return np.where(np.abs(y) < EPSILON, 0.0, x / y)

def deviatoric_tensor(A, dim=3):
    return A - 1. / dim * np.trace(A) * np.eye(dim)


# ----------------------------------------------------------------------
# Closed-form 3x3 determinant and inverse.
#
# np.linalg.det / np.linalg.inv route through generic LAPACK paths under XLA
# and pivot internally.  For 3x3 matrices the analytic Sarrus / cofactor
# formulas are typically 5-10x faster because they avoid LAPACK dispatch and
# generate a much smaller XLA HLO graph.
#
# Per cell-quad-point Newton iteration, return_map calls 6 of these
# (det F_old, det Fact, det F0, det f, det F, inv F_old).  Replacing them
# with closed-form is a pure win.
# ----------------------------------------------------------------------

def det3(M):
    return (M[0, 0] * (M[1, 1] * M[2, 2] - M[1, 2] * M[2, 1])
          - M[0, 1] * (M[1, 0] * M[2, 2] - M[1, 2] * M[2, 0])
          + M[0, 2] * (M[1, 0] * M[2, 1] - M[1, 1] * M[2, 0]))


def inv3(M):
    """Closed-form inverse of a 3x3 matrix.  Caller ensures det != 0."""
    d = det3(M)
    inv_d = 1.0 / d
    cof = np.array([
        [M[1, 1] * M[2, 2] - M[1, 2] * M[2, 1],
         M[0, 2] * M[2, 1] - M[0, 1] * M[2, 2],
         M[0, 1] * M[1, 2] - M[0, 2] * M[1, 1]],
        [M[1, 2] * M[2, 0] - M[1, 0] * M[2, 2],
         M[0, 0] * M[2, 2] - M[0, 2] * M[2, 0],
         M[0, 2] * M[1, 0] - M[0, 0] * M[1, 2]],
        [M[1, 0] * M[2, 1] - M[1, 1] * M[2, 0],
         M[0, 1] * M[2, 0] - M[0, 0] * M[2, 1],
         M[0, 0] * M[1, 1] - M[0, 1] * M[1, 0]],
    ])
    return cof * inv_d


def symmetric_eigh_perturbed(A):
    perturb = np.diag(np.array([1.17e-6, 3.7e-9, 8.27e-7]))
    vals, vecs = np.linalg.eigh(A + perturb)
    vals = np.maximum(vals, 1e-12)
    return vals, vecs

def build_axisymmetric_expmap_plasticity(E, nu, sig0, hardening_fn, hardening_prime_fn=None):
    dim = 3
    K = E / (3. * (1. - 2. * nu))
    G = E / (2. * (1. + nu))
    sqrt23 = np.sqrt(2.0 / 3.0)

    def newton_expmap(eps_e_dev_norm, alpha_old):
        def implicit_residual(d_gamma):
            alpha_eval = alpha_old + sqrt23 * d_gamma
            return (2. * G * eps_e_dev_norm
                    - 2. * G * d_gamma
                    - sqrt23 * (sig0 + hardening_fn(alpha_eval)))

        if hardening_prime_fn is not None:
            def implicit_grad(d_gamma):
                alpha_eval = alpha_old + sqrt23 * d_gamma
                return -2. * G - (2. / 3.) * hardening_prime_fn(alpha_eval)
        else:
            implicit_grad = jax.grad(implicit_residual)

        def body_fun(carry, _):
            d_gamma, converged = carry
            res = implicit_residual(d_gamma)
            res_grad = implicit_grad(d_gamma)
            d_gamma_u = lax.cond(
                converged,
                lambda d: d,
                lambda d: d - safe_divide(res, res_grad),
                d_gamma
            )
            d_gamma_u = np.where(d_gamma_u < 0., 0., d_gamma_u)
            converged_updated = np.abs(implicit_residual(d_gamma_u)) < 1.e-6
            return (d_gamma_u, converged_updated), None

        carry_init = (0.0, False)
        # (d_gamma_final, _), _ = lax.scan(body_fun, carry_init, None, length=15)
        (d_gamma_final, _), _ = lax.scan(body_fun, carry_init, None, length=30)
        return d_gamma_final

    def plastic_update(F, be_bar_trial, alpha_old):
        lam_sq, V = symmetric_eigh_perturbed(be_bar_trial)
        eps_e = 0.5 * np.log(lam_sq)
        eps_e_dev = eps_e - np.mean(eps_e)
        eps_e_dev_norm = np.linalg.norm(eps_e_dev)

        sig_flow = sig0 + hardening_fn(alpha_old)
        yield_f = 2. * G * eps_e_dev_norm - sqrt23 * sig_flow

        Delta_gamma = np.where(
            yield_f > 0.,
            newton_expmap(eps_e_dev_norm, alpha_old),
            0.
        )

        direction = np.where(
            eps_e_dev_norm > 1e-10,
            eps_e_dev / eps_e_dev_norm,
            np.zeros(dim)
        )
        eps_e_dev_new = eps_e_dev - Delta_gamma * direction
        eps_e_new = eps_e_dev_new + np.mean(eps_e)
        alpha = alpha_old + sqrt23 * Delta_gamma

        be_bar_new = V @ np.diag(np.exp(2. * eps_e_new)) @ V.T
        s_new = V @ np.diag(2. * G * eps_e_dev_new) @ V.T
        J = np.linalg.det(F)
        tau = 0.5 * K * (J ** 2 - 1.) * np.eye(dim) + s_new
        be_new = be_bar_new * (J ** (2. / 3.))
        return Delta_gamma, tau, be_new, alpha, be_bar_new

    return plastic_update

class Plasticity(Problem):
    """
    Axisymmetric plasticity problem with robust F-bar formulation.
    
    The F-bar method alleviates volumetric locking by replacing the volumetric 
    part of the deformation gradient with an element-averaged value.
    
    For axisymmetric problems, the 3D deformation gradient is:
        F = | F_rr  F_rz  0     |
            | F_zr  F_zz  0     |
            | 0     0     F_θθ  |
    
    where F_θθ = (r + u_r) / r is the hoop deformation.
    
    The F-bar formulation uses:
        F_bar = (J_avg / J_qp)^(1/3) * F
    
    where J_avg is computed via proper volume-weighted integration over the element.
    """
    
    def corrected_normals(self, ref_pt,sol):
        """
        Aligns face normals to consistently point away from a reference point.
        Args:
            points (jax.numpy.ndarray): Array of vertex coordinates, shape (num_points, 3).
            sol (jax.numpy.ndarray): The solution array containing nodal values, shape (num_nodes, dim).
            ref_pt (jax.numpy.ndarray): The reference point coordinate, shape (3,).

        Returns:
            A new array of corrected, consistently outward-pointing normal vectors at quadrature points.
        """
        sol_quad_surface = self.fes[0].convert_from_dof_to_face_quad(sol, self.boundary_inds_list[0])
        normal_vectors_qp = self.fes[0].get_physical_surface_norm(sol_quad_surface, self.boundary_inds_list[0])

        # We also need the location of these points for plotting.
        # Shape: (num_faces, num_face_quads, dim)
        boundary_inds =self.boundary_inds_list[0]
        quad_points_qp = self.fes[0].get_physical_surface_quad_points(boundary_inds, sol)
        # 1. Calculate the center of each face
        # Align ALL Normals with a Reference Point ---
        # Reference point, can be the centroid of the whole mesh for convex bodies
        # ref_pt = np.mean(self.mesh[0].points, axis=0)  # OR USE A SPECIFIC POINT
        # Reshape the arrays to be a simple list of points and vectors for easier processing
        # New shape: (total_num_quad_points, dim)
        all_quad_points = quad_points_qp.reshape(-1, self.dim)
        all_normals = normal_vectors_qp.reshape(-1, self.dim)
        # Create a vector from each quadrature point to the reference point
        vec_to_ref = ref_pt - all_quad_points
        # Compute the dot product for every single normal vector
        dot_products = np.einsum('ij,ij->i', all_normals, vec_to_ref)
        # Determine which normals to flip. For outward-pointing normals, the dot product
        # with the vector pointing to the centroid should be negative. Flip if positive.
        flip_multiplier = np.where(dot_products > 0., -1.0, 1.0)
        # Apply the flip to all inconsistent normals at once
        # The [:, np.newaxis] is used for correct broadcasting.
        corrected_normals_flat = all_normals * flip_multiplier[:, np.newaxis]
        # Reshape the corrected normals and points back to their per-face structure
        corrected_normals_qp = corrected_normals_flat.reshape(quad_points_qp.shape)
        return corrected_normals_qp

    def custom_init(self):
        """
        Initialize the axisymmetric plasticity problem with F-bar formulation.
        
        The F-bar method alleviates volumetric locking in nearly incompressible
        materials by using element-averaged volumetric deformation.
        
        Parameters
        ----------
        E : float
            Young's modulus
        sig0 : float
            Initial yield stress
        Q : float
            Isotropic hardening saturation value
        b : float
            Isotropic hardening rate parameter
        """
        self.fe = self.fes[0]
        
        # Initialize deformation gradient history (3x3 for axisymmetric)
        self.F_old = np.repeat(np.repeat(np.eye(self.dim+1)[None, None, :, :], len(self.fe.cells), axis=0),
                                self.fe.num_quads, axis=1)
        
        # Initialize elastic left Cauchy-Green tensor
        self.Be_old = np.array(self.F_old)
        
        # Initialize plastic equivalent strain
        self.alpha_old = np.zeros((len(self.fe.cells), self.fe.num_quads))
        
        # Creep internal variable (equivalent creep strain)
        self.alpha_cr_old = np.zeros((len(self.fe.cells), self.fe.num_quads))
        
        # Shape function gradients at element center (for F-bar computation)
        self.shape_grads_center = self.fe.shape_grads_center
        self.cell_sol = np.zeros_like(self.shape_grads_center)
        
        # Flexible element indices (all elements by default)
        self.fe.flex_inds = np.arange(len(self.fe.cells))
        
        # Material parameter array (for heterogeneous materials)
        full_params = np.ones((self.fe.num_cells, 4))
        self.thetas = np.repeat(full_params[:, None, :], self.fe.num_quads, axis=1)
        
        # Physical quadrature point coordinates
        physical_quad = self.physical_quad_points
        physical_quad = physical_quad[:, [0, 1, 2, 3], :]
        # Store radial coordinate for axisymmetric integration
        physical_quad[:, :, 1] = physical_quad[:, :, 0]
        self.physical_quad = np.array(physical_quad)
        
        # Shape function values at quadrature points
        self.shape_vals = np.repeat(self.fes[0].shape_vals[None, :, :, ], len(self.fe.cells), axis=0)
        
        # Jacobian times quadrature weight
        self.JxW_md = np.repeat(self.JxW, self.fe.num_quads, axis=1)
        
        # Element center coordinates (for F-bar reference)
        self.cell_center = self.fe.physical_center
        
        # Time step initialization
        dt = 1.0
        dt_full = np.ones((self.fe.num_cells, 1))
        dt_full = dt_full.at[self.fe.flex_inds].set(dt)
        self.dt = np.repeat(dt_full, self.fe.num_quads, axis=1)

        # Collect all internal variables
        self.internal_vars = jax.device_put([self.F_old, self.Be_old, self.alpha_old, self.shape_grads_center, self.cell_sol,
                              self.physical_quad, self.shape_vals, self.JxW_md,
                              self.cell_center, self.alpha_cr_old, self.thetas, self.dt])

        # ---------------------------------------------------------------
        # JIT-compiled map cache  (populated lazily on first call)
        # Keys: None until _build_compiled_maps() is called.
        # ---------------------------------------------------------------
        self._compiled_tensor_map        = None
        self._compiled_update_int_vars   = None
        self._compiled_cauchy_stress     = None
        self._compiled_lagrangian_strain = None
        self._compiled_log_strain        = None
        self._compiled_von_mises         = None

        # Reference-config surface normal cache (constant across all steps)
        self._cached_ref_normals = None

        # Reference-config surface quad-point cache (also constant for fixed mesh)
        self._cached_ref_surface_quad = None

        # thetas parameter cache (avoid rebuild when rho unchanged)
        self._cached_rho_key   = None   # stores last rho as bytes for fast comparison
        self._cached_thetas    = None

        # Per-time-step cache for set_params: skips heavy work when called multiple
        # times with identical (scale, dt, int_vars, rho_ini) -- which happens at
        # every Newton iteration within one time step.
        self._set_params_step_key  = None
        self._cached_step_int_vars = None
        self._cached_step_scales   = None
        

    def _build_compiled_maps(self):
        """
        Build and JIT-compile all vmapped maps ONCE per problem lifetime.

        get_maps() constructs closures that close over self.param (fixed bounds)
        and the per-element theta / dt internal variables (passed as arguments).
        The closure shape/structure is therefore static; only the *values* of the
        internal-variable arguments change between calls.  JAX will retrace only
        when the *abstract* shapes change, which they do not here.

        Calling get_maps() + jax.jit(jax.vmap(...)) on every forward step is the
        single largest source of unnecessary compilation overhead in the original
        code.  This method is called once, lazily, the first time any compiled map
        is needed, and thereafter the cached handles are reused.
        """
        (tensor_map, update_int_vars_map, cauchy_map,
         lagr_map, log_map) = self.get_maps()

        self._compiled_tensor_map        = jax.jit(jax.vmap(jax.vmap(tensor_map)))
        self._compiled_update_int_vars   = jax.jit(jax.vmap(jax.vmap(update_int_vars_map)))
        self._compiled_cauchy_stress     = jax.jit(jax.vmap(jax.vmap(cauchy_map)))
        self._compiled_lagrangian_strain = jax.jit(jax.vmap(jax.vmap(lagr_map)))
        self._compiled_log_strain        = jax.jit(jax.vmap(jax.vmap(log_map)))

        # von-Mises operates on (num_cells, num_quads, 3, 3) Cauchy stress arrays.
        def _von_mises_single(sigma):
            return np.sqrt(0.5 * (
                (sigma[0, 0] - sigma[1, 1]) ** 2 +
                (sigma[1, 1] - sigma[2, 2]) ** 2 +
                (sigma[2, 2] - sigma[0, 0]) ** 2) +
                3. * (sigma[0, 1] ** 2 + sigma[1, 2] ** 2 + sigma[2, 0] ** 2))
        self._compiled_von_mises = jax.jit(jax.vmap(_von_mises_single))

    def _ensure_compiled_maps(self):
        """Lazy initializer – safe to call repeatedly (no-op after first build)."""
        if self._compiled_tensor_map is None:
            self._build_compiled_maps()

    # ------------------------------------------------------------------ #
    #   Helpers for the thetas / reference-normal caches                   #
    # ------------------------------------------------------------------ #

    def _build_thetas(self, rho_ini):
        """Return the (num_cells, num_quads, num_params) theta array, using cache.

        Cache key is `id(rho_ini)` (object identity), not the value bytes.
        This is tracer-safe -- under jax.vjp / jax.grad, `rho_ini` becomes an
        abstract Tracer and `onp.asarray(tracer).tobytes()` raises
        TracerArrayConversionError.  Using id() works for both concrete and
        traced arrays.

        In practice this is also what we want: the optimizer rebuilds rho_ini
        fresh on every J_total call (via _assemble_rho), so its Python id is
        unique per optimizer step.  Within one step, the same rho_ini object
        is passed unchanged through every Newton iteration -- exactly the
        scenario the cache exists to optimize.
        """
        rho_key = id(rho_ini)
        if rho_key != self._cached_rho_key:
            full_params = np.ones((self.fe.num_cells, len(rho_ini)))
            full_params = full_params.at[self.fe.flex_inds].set(rho_ini)
            self._cached_thetas  = np.repeat(full_params[:, None, :], self.fe.num_quads, axis=1)
            self._cached_rho_key = rho_key
        return self._cached_thetas

    def _get_ref_normals(self, ref_pt, sol_shape):
        """Return reference-config surface normals, computing them at most once."""
        if self._cached_ref_normals is None:
            sol_ref = np.zeros(sol_shape)
            self._cached_ref_normals = self.corrected_normals(ref_pt, sol_ref)
        return self._cached_ref_normals

    def _get_ref_surface_quad(self, sol_shape):
        """
        Return reference-config surface quadrature points (with the radial-coord
        re-broadcast applied), computed at most once.

        The original code recomputed `physical_surface_quad_points` from the
        current `sol` on every Newton iteration.  For axisymmetric pressure
        boundary integration, the kernel only consumes the radial coordinate
        (slot 0, which is also written into slot 1).  Using the reference-config
        version is correct for total-Lagrangian surface integrals and saves
        one face-quad-points computation per Newton step.
        """
        if self._cached_ref_surface_quad is None:
            sol_ref = np.zeros(sol_shape)
            pq = self.fes[0].get_physical_surface_quad_points(self.boundary_inds_list[0], sol_ref)
            pq = pq.at[:, :, 1].set(pq[:, :, 0])
            self._cached_ref_surface_quad = pq
        return self._cached_ref_surface_quad

    def get_surface_maps(self):
        def surface_map(u, point, scale, norm_vec, physical_surface_quad_points):
            curr_pressure = scale * self.pressure_mag
            # For axisymmetric: pressure acts in direction of normal vector
            # Don't multiply by physical_surface_quad_points here - that's done in surface_kernel
            return np.array(curr_pressure * norm_vec)
        return [surface_map]

    def get_surface_kernel(self, surface_map):
        def surface_kernel(cell_sol_flat, x, face_shape_vals, face_shape_grads, face_nanson_scale,
                        *cell_internal_vars_surface):
            cell_sol_list = self.unflatten_fn_dof(cell_sol_flat)
            cell_sol = cell_sol_list[0]
            face_shape_vals = face_shape_vals[:, :self.fes[0].num_nodes]
            face_shape_grads = face_shape_grads[:, :self.fes[0].num_nodes, :]
            face_nanson_scale = face_nanson_scale[0]

            # Interpolate u and grad(u)
            u = np.sum(cell_sol[None, :, :] * face_shape_vals[:, :, None], axis=1)
            # u_grad = np.einsum('qnd,ndv->qdv', face_shape_grads, cell_sol)  # ∂u/∂X
            u_grad = jax.vmap(lambda dNdx: dNdx.T @ cell_sol)(face_shape_grads)
            u_grad = np.transpose(u_grad, axes=(0, 2, 1))  # now shape is (num_quads, vec, dim)
            r_coords = x[:, 0] # x: (num_face_quads, 2)
            u_r = u[:, 0] # u: (num_face_quads, 2)

            # Build 3D deformation gradient F
            F = np.zeros((u.shape[0], 3, 3))
            F = F.at[:, :2, :2].set(u_grad + np.eye(2))  # Top-left 2x2 block
            F = F.at[:, 2, 2].set((r_coords + u_r) / r_coords)  # F_theta_theta
            J = np.linalg.det(F)
            FinvT = jax.vmap(lambda A: np.linalg.inv(A).T)(F)

            # Estimate reference normal vector at each quad point
            # n0 = self.get_reference_normals(x)  # (num_face_quads, dim)
            # Call 2D surface_map(u, x, scale, norm_vec)
            pressure_2d = jax.vmap(surface_map)(u, x, *cell_internal_vars_surface)  # (num_face_quads, 2)
                # Promote to 3D vector: [pr, pz, 0]
            pressure_3d = np.concatenate([
                pressure_2d[:, :2], np.zeros((pressure_2d.shape[0], 1))], axis=1)
            
            # Convert to traction in reference config: t0 = -J * FinvT @ pressure_vec
            t0_3d = J[:, None] * jax.vmap(np.matmul)(FinvT, pressure_3d)  # (num_face_quads, dim) -
            # Convert back to 2D: [r, z] → drop θ
            t0 = t0_3d[:, :2]  # (num_face_quads, 2)
            # Project to shape functions (TL surface integral)
            # CRITICAL: For axisymmetric formulation, must multiply by r coordinate
            val = np.sum(face_shape_vals[:, :, None] * t0[:, None, :] * face_nanson_scale[:, None, None] * r_coords[:, None, None], axis=0)
            # jax.debug.print("rcord,{}", r_coords[:, None, None])
            # jax.debug.print('val,{}',np.shape(val))
            return jax.flatten_util.ravel_pytree(val)[0]
        return surface_kernel

    def get_laplace_kernel(self, tensor_map):
        def laplace_kernel(cell_sol_flat,cell_shape_grads, cell_v_grads_JxW, *cell_internal_vars):
            # cell_sol_flat: (num_nodes*vec + ...,)
            # cell_sol_list: [(num_nodes, vec), ...]
            # cell_shape_grads: (num_quads, num_nodes + ..., dim)
            # cell_v_grads_JxW: (num_quads, num_nodes + ..., 1, dim)

            cell_sol_list = self.unflatten_fn_dof(cell_sol_flat)
            cell_shape_grads = cell_shape_grads[:, :self.fes[0].num_nodes, :]
            cell_sol = cell_sol_list[0]
            cell_v_grads_JxW = cell_v_grads_JxW[:, :self.fes[0].num_nodes, :, :]
            vec = self.fes[0].vec

            # (1, num_nodes, vec, 1) * (num_quads, num_nodes, 1, dim) -> (num_quads, num_nodes, vec, dim)
            u_grads = cell_sol[None, :, :, None] * cell_shape_grads[:, :, None, :]
            u_grads = np.sum(u_grads, axis=1)  # (num_quads, vec, dim)


            # a1,a2,a3,a4,a5,a6=cell_internal_vars JxW =a8
            a1, a2, a3, a4, a5, a6,a7,a8,a9,a10,a11,a12 = cell_internal_vars
            # Here------>
            # a4 = center gradient
            # a5 = cell solution
            # a6 = radial distance to the Gauss points
            # a7 = shape functions
            # a8 = JxW
            # a9 = cell center
            # a10 = theta
            # cell_internal_vars_updated = (a1,a2,a3,a4,u_grads_ceter_reshape, a6)
            ## cell_sol/a6 ==> Consider shape function in this expression --> u = u1*N1+ u2*N2+ u3*N3+ u4*N4--<
            cell_sol_shape = self.fes[0].shape_vals[:,:,None]*cell_sol[None,:,:]
            sol_shape = np.sum(cell_sol_shape, axis=1)  # +val_sigma3
        #    jax.debug.print('cell_sol_shape,{},{}', self.fes[0].shape_vals[:,:,None],a6)
            # jax.debug.print('sol_shape_by_r,{}', sol_shape/a6)
            # calculate u at Gauss points
            a5_update = cell_sol[None,:,:]*np.ones_like(a4)
            # cell_internal_vars_updated = (a1, a2, a3, a4, cell_sol,sol_shape/a6,center_sol/center_r,a8,a9) #a6
            cell_internal_vars_updated = (a1, a2, a3, a4, a5_update, a6, a7, a8, a9,a10,a11,a12)  # a6

            physical_quad = a6.reshape(-1, 1, self.dim)
            # logger.debug(f"u_grads, u_grads = {u_grads}")
            u_grads_reshape = u_grads.reshape(-1, vec, self.dim)  # (num_quads, vec, dim)
            # (num_quads, vec, dim)
            stress = jax.vmap(tensor_map)(u_grads_reshape, *cell_internal_vars_updated).reshape((4,3,3))
            stress_2D = stress[:,0:2,0:2]
            stress_33 = stress[:,2,2]
            sigma_by_r = stress_33/a6[:,0]
            shape_fun = a7
            JxW = a8
            # (num_quads, num_nodes, vec, dim) -> (num_nodes, vec) * physical_quad[:, None, :, :]
            # val = np.sum(stress[:, None, :, :] * cell_v_grads_JxW, axis=(0, -1))# multiply physical_quad r coord here
            val_2D = stress_2D[:, None, :, :] * cell_v_grads_JxW
            val_sigma3 = np.zeros_like(val_2D)
            # jax.debug.print('stress,{}', stress)
            val_sigma3 = val_sigma3.at[:, :, 0, 0].set(shape_fun * JxW * sigma_by_r[:, None])
            
            val = np.sum((val_2D+val_sigma3)*physical_quad[:, None, :, :], axis=(0, -1))  # +val_sigma3
            # jax.debug.print('val,{}', val)
            # val = np.sum(stress[:, None, :, :] * u_grads_axi , axis=(0, -1)) #*physical_quad[:, None, :, :]
            val = jax.flatten_util.ravel_pytree(val)[0]  # (num_nodes*vec + ...,)
            return val
        return laplace_kernel

    def get_tensor_map(self):
        # Return the raw (un-vmapped) tensor_map callable as before so that the
        # FEM assembly framework can vmap it itself.  We still avoid rebuilding the
        # closure by delegating to get_maps() only if a fresh closure is genuinely
        # needed.  The compiled *vmapped* version is accessed via _compiled_tensor_map.
        tensor_map, _, _, _, _ = self.get_maps()
        return tensor_map

    def get_maps(self):
        # Paper used for properties
        # https://www.sciencedirect.com/science/article/pii/S1359645420309939

        def get_partial_tensor_map(F_old, be_old, alpha_old, shape_grads_center, 
                                   cell_sol,r_gauss,shape_val,JxW,cell_center,
                                   alpha_cr_old,theta,dt):
            """
            Robust F-bar formulation for axisymmetric elements.
            
            The F-bar method uses element-averaged volumetric deformation:
                F_bar = (J_avg / J)^(1/3) * F
            
            For axisymmetric problems, proper volume weighting includes 2*pi*r factor.
            """
            # Calculate gradient at the center
            u_grads_center = cell_sol[:, :, None] * shape_grads_center[:, None, :]
            ugrad0 = np.sum(u_grads_center, axis=0)  # (vec, dim)

            # Find the center of the cell and repeat it at 4 quad points
            center_rad = cell_center[0]
            center_sol = np.mean(cell_sol, axis=0)
            center_r = center_sol / center_rad

            # (Shape function*u)/r_gauss for hoop strain
            cell_sol_shape = shape_val[:, None] * cell_sol
            sol_shape = np.sum(cell_sol_shape, axis=0)
            NuByR = sol_shape / r_gauss

            # jax.debug.print("theta,{}", theta)
            theta_mod = (theta + 1) / 2
            E = self.param[0][0] + theta_mod[0] * self.param[0][1]
            sig0 = self.param[1][0] + theta_mod[1] * self.param[1][1]
            Q = self.param[2][0] + theta_mod[2] * self.param[2][1]
            b = self.param[3][0] + theta_mod[3] * self.param[3][1]
            
            # Creep parameters (Norton-Bailey)
            A = 0
            n = self.param[5][0] + theta_mod[5] * self.param[5][1]
            
            nu = 0.3
            # jax.debug.print("theta,{}", theta)
            K = E / (3. * (1. - 2. * nu))
            G = E / (2. * (1. + nu))

            def first_PK_stress(u_grad):
                F, _, _, tau, _, _ = return_map(u_grad)
                # Closed-form 3x3 inverse — used per cell-quad-point per Newton iter.
                P = tau @ inv3(F).T
                return P

            def update_int_vars(u_grad):
                F, be_bar, alpha, _, _, alpha_cr = return_map(u_grad)
                return F, be_bar, alpha, shape_grads_center, cell_sol, r_gauss, shape_val, JxW, cell_center, alpha_cr, theta, dt

            def compute_cauchy_stress(u_grad):
                F, _, _, tau, _, _ = return_map(u_grad)
                J = det3(F)              # closed-form 3x3 det
                sigma = (1. / J) * tau
                return sigma

            def compute_lagrangian_strain(u_grad):
                u_grad_axi = np.zeros((3, 3))
                u_grad_axi = u_grad_axi.at[0:2, 0:2].set(u_grad)
                u_grad_axi = u_grad_axi.at[2, 2].set(NuByR[0])
                F = u_grad_axi + np.eye(self.dim + 1)
                strain = 0.5 * (F @ F.T - np.eye(self.dim + 1))
                return strain

            def compute_logarithmic_strain(u_grad):
                # Previously called return_map(u_grad) just to get F, which forced
                # JAX to trace the entire plasticity-creep return-mapping closure
                # (solve_creep_implicit + solve_coupled + a 18-iteration scan).
                # Compute F directly from u_grad with the same F-bar correction
                # used inside return_map, completely sidestepping the inner solves.
                u_grad_axi = np.zeros((3, 3))
                u_grad_axi = u_grad_axi.at[0:2, 0:2].set(u_grad)
                u_grad_axi = u_grad_axi.at[2, 2].set(NuByR[0])
                Fact = u_grad_axi + np.eye(self.dim + 1)

                # F-bar correction (mirrors return_map)
                ugrad0_axi = np.zeros((3, 3))
                ugrad0_axi = ugrad0_axi.at[0:2, 0:2].set(ugrad0)
                ugrad0_axi = ugrad0_axi.at[2, 2].set(center_r[0])
                F0 = ugrad0_axi + np.eye(self.dim + 1)

                J_qp  = det3(Fact)
                J_avg = det3(F0)
                J_qp_safe  = np.where(np.abs(J_qp)  > 1e-14, J_qp,  1e-14 * np.sign(J_qp + 1e-20))
                J_avg_safe = np.where(np.abs(J_avg) > 1e-14, J_avg, 1e-14)
                J_ratio    = J_avg_safe / J_qp_safe
                fbar_scale = np.sign(J_ratio) * (np.abs(J_ratio)) ** (1.0 / 3.0)
                F = fbar_scale * Fact

                # Logarithmic strain via eigendecomposition of F F^T
                e_val, e_vec = np.linalg.eigh(F @ F.T)
                e_val_safe   = np.maximum(e_val, 1e-14)  # guard log of tiny eigenvalues
                log_strain   = 0.5 * (e_vec @ np.diag(np.log(e_val_safe)) @ e_vec.T)
                ldg  = np.eye(self.dim + 1) * log_strain
                loff = 2. * (log_strain - ldg)
                return ldg + loff

            def get_tau(F, be_bar):
                J = np.linalg.det(F)
                tau = 0.5 * K * (J ** 2 - 1.) * np.eye(self.dim + 1) + G * deviatoric_tensor(be_bar, self.dim + 1)
                return tau

            def K_fun(a):
                nonlinear = Q * (1. - np.exp(-b * a))
                return nonlinear

            def dK_dalpha_fun(a):
                return Q * b * np.exp(-b * a)

            # NOTE: build_axisymmetric_expmap_plasticity was previously called here to
            # build a `plasticity_expmap` closure, but V10's return_map uses the manual
            # solve_creep_implicit + solve_coupled path and never invokes plasticity_expmap.
            # JAX still traced through the unused closure (which contains its own
            # linalg.eigh + Newton solve), bloating compilation.  Removed.

            def compute_element_average_jacobian(cell_sol, shape_grads_center, shape_val, r_gauss, JxW, cell_center):
                """
                Compute the element-averaged Jacobian using volume-weighted integration.
                
                For axisymmetric elements, the volume integral is:
                    J_avg = (∫ J * r * dA) / (∫ r * dA)
                
                where the 2*pi factor cancels out.
                
                This method uses proper Gauss quadrature for numerical integration.
                
                Returns:
                    J_avg: Element-averaged Jacobian
                    J_center: Jacobian at element center (for reference)
                """
                # Compute gradient at element center
                u_grads_center_local = cell_sol[:, :, None] * shape_grads_center[:, None, :]
                ugrad0_local = np.sum(u_grads_center_local, axis=0)
                
                # Center coordinates
                center_rad_local = cell_center[0]
                center_sol_local = np.mean(cell_sol, axis=0)
                center_r_local = center_sol_local / center_rad_local
                
                # F at center
                ugrad0_axi = np.zeros((3, 3))
                ugrad0_axi = ugrad0_axi.at[0:2, 0:2].set(ugrad0_local)
                ugrad0_axi = ugrad0_axi.at[2, 2].set(center_r_local[0])
                F_center = ugrad0_axi + np.eye(3)
                J_center = np.linalg.det(F_center)
                
                return J_center

            def return_map(u_grad):
                """
                Combined plasticity-creep return mapping algorithm with robust F-bar formulation.
                
                F-bar Method for Axisymmetric Elements:
                ======================================
                The standard F-bar method modifies the deformation gradient to use
                element-averaged volumetric deformation:
                
                    F_bar = (J_avg / J)^(1/3) * F
                
                This alleviates volumetric locking by ensuring that the volumetric
                response is represented uniformly across the element.
                
                For axisymmetric problems, we must properly account for:
                1. The 3x3 deformation gradient with hoop component F_θθ
                2. Volume weighting with radial coordinate r
                3. Proper treatment near the axis of symmetry (r → 0)
                
                Algorithm (coupled):
                1. Compute trial elastic state with F-bar
                2. Solve creep implicitly (fix plasticity)
                3. Check yield condition
                4. If yield, solve coupled system
                5. Return updated stress and internal variables
                """
                
                # Elastic predictor
                # Old isochoric left Cauchy-Green tensor
                # Use closed-form det3 (5-10x faster than np.linalg.det for 3x3).
                be_bar_old = (det3(F_old) ** (-2.0 / 3.0)) * be_old

                # ================================================================
                # ROBUST F-BAR FORMULATION FOR AXISYMMETRIC ELEMENTS
                # ================================================================
                
                # Step 1: Compute deformation gradient at current quadrature point
                u_grad_axi = np.zeros((3, 3))
                u_grad_axi = u_grad_axi.at[0:2, 0:2].set(u_grad)
                u_grad_axi = u_grad_axi.at[2, 2].set(NuByR[0])
                Fact = u_grad_axi + np.eye(self.dim + 1)
                
                # Step 2: Compute Jacobian at quadrature point (closed-form 3x3 det)
                J_qp = det3(Fact)
                
                # Step 3: Compute element-averaged Jacobian (using center point approximation)
                # For axisymmetric F-bar, compute F at element center using proper interpolation
                # The displacement at center uses shape function values at (ξ=0, η=0) which are all 0.25
                # So u_center = 0.25 * (u1 + u2 + u3 + u4) = mean(cell_sol)
                ugrad0_axi = np.zeros((3, 3))
                ugrad0_axi = ugrad0_axi.at[0:2, 0:2].set(ugrad0)
                ugrad0_axi = ugrad0_axi.at[2, 2].set(center_r[0])
                F0 = ugrad0_axi + np.eye(self.dim + 1)
                J_avg = det3(F0)
                
                # Step 4: Compute F-bar scaling factor with numerical safeguards
                # Proper F-bar formulation: F_bar = (J_avg/J)^(1/3) * F
                # This modifies only the volumetric part while preserving the deviatoric part
                
                # Numerical safeguards for Jacobians
                J_qp_safe = np.where(np.abs(J_qp) > 1e-14, J_qp, 1e-14 * np.sign(J_qp + 1e-20))
                J_avg_safe = np.where(np.abs(J_avg) > 1e-14, J_avg, 1e-14)
                
                # Compute scaling factor: (J_avg / J_qp)^(1/3)
                J_ratio = J_avg_safe / J_qp_safe
                
                # For robustness, use sign-preserving cube root
                # This handles both positive and negative Jacobians correctly
                fbar_scale = np.sign(J_ratio) * (np.abs(J_ratio)) ** (1.0 / 3.0)
                
                # Apply F-bar modification
                # F_bar = (J_avg / J)^(1/3) * F
                F = fbar_scale * Fact
                # F =  Fact
                
                # ================================================================
                # END F-BAR FORMULATION
                # ================================================================
                
                # Closed-form 3x3 inverse (avoids LAPACK pivoting + dispatch).
                F_old_inv = inv3(F_old)
                
                # Relative deformation gradient
                f = F @ F_old_inv
                
                # Isochoric part of relative deformation gradient (closed-form det)
                J_f = det3(f)
                J_f_safe = np.where(np.abs(J_f) > 1e-14, J_f, 1e-14)
                f_bar = (np.abs(J_f_safe) ** (-1. / 3.)) * f
                
                # Trial elastic isochoric left Cauchy-Green tensor
                be_bar_trial = f_bar @ be_bar_old @ f_bar.T

                # ====================================================================
                # PRINCIPAL-STRAIN EXP-MAP RETURN MAPPING
                #
                # Decompose b_e_bar^trial into its spectral form once, then carry the
                # whole return mapping (yield surface, creep, reconstruction) in
                # principal-logarithmic-strain space.  This makes the final state
                # match Simo's exp-map flow rule exactly:
                #
                #     b_e^{n+1} = exp(-2*(gamma_p+gamma_cr)*N) : b_e^trial
                #
                # so the response is independent of step size to the inner-Newton
                # tolerance, removing the O(strain^2) error of the small-strain
                # radial-return reconstruction (s_new / G + Ie_bar * I).
                #
                # Algebra: along the fixed flow direction N_a = s_trial_a / ||s_trial||,
                # the corrected principal logarithmic strains are
                #     eps^{n+1}_a = eps^trial_a - (gamma_p + gamma_cr) * N_a
                # and the corrected principal deviatoric Kirchhoff stresses are
                #     s^{n+1}_a   = 2*G * eps^{n+1}_a
                # so the scalar reduction in ||s|| is exactly 2*G*(gamma_p + gamma_cr)
                # -- structurally identical to the previous formulation but with G
                # in place of the small-strain trace-averaged G_bar.
                # ====================================================================

                # symmetric_eigh_perturbed adds tiny off-diagonal perturbations that
                # break degeneracy on near-isotropic states (early-history elastic
                # increments).  This is essential for clean adjoint gradients
                # through linalg.eigh; pure isotropy makes eigh's gradient singular.
                lam_sq, eig_vecs = symmetric_eigh_perturbed(be_bar_trial)
                lam_sq = np.maximum(lam_sq, 1e-14)             # safety floor
                eps_trial_a = 0.5 * np.log(lam_sq)             # principal log-strains
                # Detrace (be_bar_trial is isochoric so tr(eps_trial) ~ 0 numerically,
                # but enforce it strictly to avoid tiny drifts from the perturbation).
                eps_trial_a = eps_trial_a - (1.0 / 3.0) * np.sum(eps_trial_a)

                # Principal trial deviatoric Kirchhoff stresses
                s_trial_a = 2.0 * G * eps_trial_a
                s_trial_norm = np.linalg.norm(s_trial_a)

                # Flow direction in eigenspace.  Use the safe-divide pattern from the
                # original code -- when ||s_trial|| is tiny the direction is irrelevant
                # because both gamma_p and gamma_cr will be ~0 anyway.
                s_trial_norm_safe = np.maximum(s_trial_norm, 1e-14)
                N_a = np.where(
                    s_trial_norm > 1e-14,
                    s_trial_a / s_trial_norm_safe,
                    np.zeros_like(s_trial_a),
                )

                sqrt23 = np.sqrt(2.0 / 3.0)
                sqrt32 = np.sqrt(3.0 / 2.0)

                # ---- Step 1: pure-creep solve assuming gamma_p = 0 -------------
                # NOTE: G replaces G_bar throughout the inner solves -- the residual
                # equations are structurally unchanged but use the actual shear
                # modulus on the principal-log-strain trial state.
                gamma_cr_creep_only, _, _ = solve_creep_implicit(
                    s_trial_norm, G, alpha_cr_old
                )

                # ---- Step 2: yield check on the post-creep stress with OLD alpha_p
                s_norm_after_creep = s_trial_norm - 2. * G * gamma_cr_creep_only
                sigma_y_old = sig0 + K_fun(alpha_old)
                yield_after_creep = s_norm_after_creep - sqrt23 * sigma_y_old
                plastic_active = yield_after_creep > 0.

                # ---- Step 3: coupled solve (only meaningful when plastic_active)
                gamma_p_coupled, gamma_cr_coupled = solve_coupled(
                    s_trial_norm, G, alpha_old, gamma_cr_creep_only
                )

                # ---- Step 4: select branch
                gamma_p  = np.where(plastic_active, gamma_p_coupled,  0.0)
                gamma_cr = np.where(plastic_active, gamma_cr_coupled, gamma_cr_creep_only)

                # Clamp non-negative (safety; Newton already enforces this)
                gamma_p  = np.maximum(gamma_p,  0.0)
                gamma_cr = np.maximum(gamma_cr, 0.0)

                # ---- Step 5: update hardening variables (slip convention)
                alpha_p  = alpha_old    + sqrt23 * gamma_p
                alpha_cr = alpha_cr_old + sqrt23 * gamma_cr

                # ---- Step 6: reconstruct b_e^{n+1} via exp-map in principal axes
                # eps^{n+1}_a = eps^trial_a - (gamma_p + gamma_cr) * N_a
                # lam^{n+1}_a = exp(eps^{n+1}_a)
                # b_e^{n+1}   = sum_a (lam^{n+1}_a)^2 * (n_a otimes n_a)
                #             = V * diag(lam^{n+1}_a^2) * V^T
                # which is exactly exp(-2*(gamma_p+gamma_cr)*N) : b_e^trial.
                eps_new_a = eps_trial_a - (gamma_p + gamma_cr) * N_a
                lam_new_sq = np.exp(2.0 * eps_new_a)           # = (exp(eps_new))^2

                # Tensor reconstruction: V @ diag(lam_new_sq) @ V^T
                be_updated_bar = (eig_vecs * lam_new_sq[None, :]) @ eig_vecs.T

                # Corrected deviatoric Kirchhoff stress (principal form is exact;
                # equivalent to s = G * dev(b_e_bar^{n+1}) but cheaper since we have
                # the principal values already).
                s_new_a = 2.0 * G * eps_new_a
                s_new   = (eig_vecs * s_new_a[None, :]) @ eig_vecs.T

                # Volumetric Kirchhoff stress (closed-form det)
                J_F = det3(F)
                tau = 0.5 * K * (J_F**2 - 1.0) * np.eye(self.dim + 1) + s_new

                # Reconstruct full be (not isochoric) for storage
                be_updated = be_updated_bar * (J_F ** (2.0 / 3.0))

                return F, be_updated, alpha_p, tau, ugrad0, alpha_cr
            
            def solve_creep_implicit(s_norm, G_eff, alpha_cr_old_val):
                """
                Implicit backward Euler solution for pure creep (no plasticity).

                NOTE on G_eff: this routine is now invoked from the principal-strain
                exp-map return map with G_eff = G (the actual shear modulus).  The
                residual algebra is unchanged from the previous trial-deviator-norm
                formulation -- only the meaning of the input shear-modulus argument
                differs.  The 2nd arg used to be the trace-averaged G_bar = G * Ie_bar
                (a small-strain artefact); it is now exact.

                gamma_cr is defined as the CREEP SLIP:

                    gamma_cr_slip = sqrt(3/2) * d_eps_cr_eq

                so the stress-norm reduction along the flow direction is
                    ||s_new|| = ||s_trial|| - 2*G_eff*gamma_cr.

                Norton-Bailey equivalent creep strain rate:
                    d_eps_cr_eq/dt = A * sigma_vm^n
                    sigma_vm = sqrt(3/2) * s_norm_current

                Residual (in slip space):
                    R = sqrt(2/3)*gamma_cr - dt * A * sigma_vm^n = 0

                Returns:
                    gamma_cr_final : converged creep slip increment
                    res_final      : final scaled residual
                    converged_flag : 1.0 if |res|<tol, else 0.0
                """
                sqrt23 = np.sqrt(2.0 / 3.0)
                sqrt32 = np.sqrt(3.0 / 2.0)
                tol = 1.0e-10

                def body_fun(state, _):
                    gamma_cr, _, conv = state

                    # Stress norm after creep correction (same formula as plasticity)
                    s_norm_current = np.maximum(s_norm - 2. * G_eff * gamma_cr, 0.0)
                    sigma_vm = sqrt32 * s_norm_current

                    # Norton-Bailey equivalent strain rate
                    creep_rate = A * np.where(sigma_vm > 0., sigma_vm ** n, 0.)

                    # Residual in slip space: sqrt(2/3)*gamma_cr = dt*A*sigma_vm^n
                    residual = sqrt23 * gamma_cr - dt * creep_rate

                    # Analytical Jacobian: dR/d(gamma_cr)
                    dsigma_vm_dgamma = sqrt32 * (-2. * G_eff)
                    dcreep_dsigma = A * n * np.where(sigma_vm > 0., sigma_vm ** (n - 1.), 0.)
                    jac = sqrt23 - dt * dcreep_dsigma * dsigma_vm_dgamma

                    delta_gamma = np.where(
                        np.abs(jac) > 1e-14,
                        -residual / jac,
                        0.,
                    )
                    delta_gamma = np.where(conv > 0.5, 0.0, delta_gamma)
                    gamma_cr_new = np.maximum(gamma_cr + delta_gamma, 0.0)

                    # Re-evaluate for convergence check
                    s_chk = np.maximum(s_norm - 2. * G_eff * gamma_cr_new, 0.0)
                    sv_chk = sqrt32 * s_chk
                    res_chk = sqrt23 * gamma_cr_new - dt * A * np.where(sv_chk > 0., sv_chk ** n, 0.)
                    conv_new = np.where(np.abs(res_chk) < tol, 1.0, conv)

                    return (gamma_cr_new, res_chk, conv_new), None

                init_state = (0.0, 1.0, 0.0)
                # 12 iterations is sufficient for the implicit Newton solve with
                # analytical Jacobian on Norton-Bailey creep (typical convergence: 3-6 iters).
                final_state, _ = jax.lax.scan(body_fun, init_state, None, length=12)
                gamma_cr_final, res_final, conv_final = final_state
                return gamma_cr_final, res_final, conv_final
            
            def solve_coupled(s_trial_norm, G_eff, alpha_p_old_val, gamma_cr_guess):
                """
                Solve fully coupled plasticity-creep system.

                NOTE on G_eff: this routine is now invoked from the principal-strain
                exp-map return map with G_eff = G (the actual shear modulus), where
                s_trial_norm is the L2 norm of the trial principal deviatoric
                Kirchhoff stresses s_a^trial = 2*G*eps_a^trial.  The 2x2 residual
                algebra is unchanged from the previous trial-deviator-norm form --
                only the meaning of the input shear-modulus differs.  G_bar (the
                small-strain trace-averaged approximation) is no longer used.

                Convention: BOTH gamma_p and gamma_cr are SLIPS (Simo exp-map sense).
                  - principal-stress correction: s^{n+1}_a = s^trial_a - 2*G*(g_p+g_cr)*N_a
                  - hardening updates:
                        alpha_p  += sqrt(2/3) * gamma_p    (equiv plastic strain)
                        alpha_cr += sqrt(2/3) * gamma_cr   (equiv creep strain)

                Residuals (scaled to O(1)):
                  R1: plastic consistency
                      s_norm - sqrt(2/3)*sigma_y(alpha_p_old + sqrt(2/3)*gamma_p) = 0
                      Scaled by sig0.
                  R2: creep evolution (backward Euler, in slip space)
                      sqrt(2/3)*gamma_cr - dt*A*sigma_vm^n = 0
                      Scaled by max(sqrt(2/3)*gamma_cr_guess, 1e-10).
                """
                sqrt23 = np.sqrt(2.0 / 3.0)
                sqrt32 = np.sqrt(3.0 / 2.0)
                tol = 1.0e-8
                
                # MODIFIED NOW
                # R1_scale = np.maximum(sig0, 1.0)
                R1_scale = np.maximum(sig0 + K_fun(alpha_p_old_val), 1.0)
                R2_scale = np.maximum(sqrt23 * gamma_cr_guess, 1.0e-10)

                def residuals(gamma_p, gamma_cr):
                    s_norm = np.maximum(
                        s_trial_norm - 2. * G_eff * (gamma_p + gamma_cr), 1e-14
                    )
                    sigma_vm = sqrt32 * s_norm
                    alpha_p_new = alpha_p_old_val + sqrt23 * gamma_p
                    sigma_y = sig0 + K_fun(alpha_p_new)

                    # R1: plastic consistency (yield surface)
                    R1 = (s_norm - sqrt23 * sigma_y) / R1_scale
                    # R2: creep evolution in slip-space convention
                    creep_term = dt * A * (sigma_vm ** n)
                    R2 = (sqrt23 * gamma_cr - creep_term) / R2_scale
                    return np.array([R1, R2]), s_norm, sigma_vm, alpha_p_new

                def jacobian(gamma_p, gamma_cr, s_norm, sigma_vm, alpha_p_new):
                    ds_norm_dg   = -2. * G_eff
                    dsigma_vm_dg = sqrt32 * ds_norm_dg          # negative
                    d_sigy_d_gp  = dK_dalpha_fun(alpha_p_new) * sqrt23

                    # dR1/dgp, dR1/dgcr
                    dR1_dgp  = (ds_norm_dg - sqrt23 * d_sigy_d_gp) / R1_scale
                    dR1_dgcr = ds_norm_dg / R1_scale

                    # dR2/dgp, dR2/dgcr
                    dcreep_dsigma_vm = dt * A * n * (sigma_vm ** (n - 1.))
                    dR2_dgp  = (-dcreep_dsigma_vm * dsigma_vm_dg) / R2_scale
                    dR2_dgcr = (sqrt23 - dcreep_dsigma_vm * dsigma_vm_dg) / R2_scale

                    return dR1_dgp, dR1_dgcr, dR2_dgp, dR2_dgcr

                def body_fun(state, _):
                    x, _, conv = state
                    gamma_p, gamma_cr = x[0], x[1]

                    R, s_norm, sigma_vm, alpha_p_new = residuals(gamma_p, gamma_cr)
                    J11, J12, J21, J22 = jacobian(
                        gamma_p, gamma_cr, s_norm, sigma_vm, alpha_p_new
                    )

                    detJ = J11 * J22 - J12 * J21
                    detJ_safe = np.where(np.abs(detJ) < 1e-14, 1.0, detJ)

                    dgp_full  = -( J22 * R[0] - J12 * R[1]) / detJ_safe
                    dgcr_full = -(-J21 * R[0] + J11 * R[1]) / detJ_safe
                    dgp_full  = np.where(np.abs(detJ) < 1e-14, 0., dgp_full)
                    dgcr_full = np.where(np.abs(detJ) < 1e-14, 0., dgcr_full)

                    R_norm_old = np.linalg.norm(R)

                    def try_step(alpha_step):
                        gp_t  = np.maximum(gamma_p  + alpha_step * dgp_full,  0.0)
                        gcr_t = np.maximum(gamma_cr + alpha_step * dgcr_full, 0.0)
                        R_t, _, _, _ = residuals(gp_t, gcr_t)
                        return gp_t, gcr_t, np.linalg.norm(R_t)

                    # Two-trial line search: first try alpha=1, fall back to 0.25 if rejected.
                    # The original 4-trial cascade (1.0, 0.5, 0.25, 0.125) cost 4 residual
                    # evaluations on every scan iteration, but in practice alpha=1 accepts
                    # ~80% of the time near convergence, and 0.25 is enough fallback for
                    # the rest.  Inside lax.scan all branches still execute, but going from
                    # 4 to 2 cuts inner-Newton residual evals roughly in half.
                    gp_a, gcr_a, Rn_a = try_step(1.0)
                    # MODIFIED NOW
                    # gp_b, gcr_b, Rn_b = try_step(0.25)
                    gp_b, gcr_b, Rn_b = try_step(0.5)
                    gp_c, gcr_c, Rn_c = try_step(0.125)

                    accept_a = Rn_a <= R_norm_old
                    accept_b = Rn_b <= R_norm_old

                    gp_new  = np.where(accept_a, gp_a,  np.where(accept_b, gp_b,  gp_c))
                    gcr_new = np.where(accept_a, gcr_a, np.where(accept_b, gcr_b, gcr_c))

                    gp_new  = np.where(conv > 0.5, gamma_p,  gp_new)
                    gcr_new = np.where(conv > 0.5, gamma_cr, gcr_new)

                    R_new, _, _, _ = residuals(gp_new, gcr_new)
                    R_norm_new = np.linalg.norm(R_new)
                    conv_new = np.where(R_norm_new < tol, 1.0, conv)

                    return (np.array([gp_new, gcr_new]), R_new, conv_new), None

                init_x = np.array([0.0, np.maximum(gamma_cr_guess, 0.0)])
                init_state = (init_x, np.array([1.0, 1.0]), 0.0)
                # 18 iterations is ample for the 2×2 coupled Newton solve with backtracking.
                # The original length=50 was conservative for lax.scan's lack of true
                # early exit; 18 provides safety margin while cutting unneeded compute.
                final_state, _ = jax.lax.scan(body_fun, init_state, None, length=18)
                final_x = final_state[0]

                return final_x[0], final_x[1]
            
            return (first_PK_stress, update_int_vars, compute_cauchy_stress, 
                   compute_lagrangian_strain, compute_logarithmic_strain)

        def tensor_map(u_grad, *params):
            first_PK_stress, _, _, _, _ = get_partial_tensor_map(*params)
            return first_PK_stress(u_grad)
        
        def update_int_vars_map(u_grad, *params):
            _, update_int_vars, _, _, _ = get_partial_tensor_map(*params)
            return update_int_vars(u_grad)
        
        def compute_cauchy_stress_map(u_grad, *params):
            _, _, compute_cauchy_stress, _, _ = get_partial_tensor_map(*params)
            return compute_cauchy_stress(u_grad)
        
        def compute_lagrangian_strain_map(u_grad, *params):
            _, _, _, compute_lagrangian_strain, _ = get_partial_tensor_map(*params)
            return compute_lagrangian_strain(u_grad)
        
        def compute_logarithmic_strain_map(u_grad, *params):
            _, _, _, _, compute_logarithmic_strain = get_partial_tensor_map(*params)
            return compute_logarithmic_strain(u_grad)
        
        return (tensor_map, update_int_vars_map, compute_cauchy_stress_map, compute_lagrangian_strain_map,
                compute_logarithmic_strain_map)

    def update_int_vars_gp(self, sol, int_vars, dt_current):
        # Use the pre-compiled, cached vmapped map instead of re-JIT-ing on every call.
        self._ensure_compiled_maps()
        vmap_update_int_vars_map = self._compiled_update_int_vars

        u_grads1 = (np.take(sol, self.fe.cells, axis=0)[:, None, :, :, None] *
                    self.fe.shape_grads[:, :, :, None, :])
        u_grads = np.sum(u_grads1, axis=2)  # (num_cells, num_quads, vec, dim)
        cell_sol = np.take(sol, self.fe.cells, axis=0)

        a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12 = int_vars
        a5_update = np.repeat(cell_sol[:, None, :], self.fe.num_quads, axis=1)

        # flex_inds is always np.arange(num_cells) so the scatter is a no-op;
        # build the (num_cells, num_quads) dt array directly.
        dt = np.full((self.fe.num_cells, self.fe.num_quads), dt_current)

        int_vars_updated = (a1, a2, a3, a4, a5_update, a6, a7, a8, a9, a10, a11, dt)
        updated_int_vars = vmap_update_int_vars_map(u_grads, *int_vars_updated)
        return updated_int_vars

    def compute_max_creep_inc(self, sol, int_vars_old, int_vars_new, dt):
        """
        Compute the maximum equivalent creep strain increment over all Gauss
        points in the mesh for the last accepted time step.

        The equivalent creep strain accumulator is internal variable index 9
        (alpha_cr, zero-indexed), i.e. int_vars[9].  The increment is:

            Delta_alpha_cr = alpha_cr_new - alpha_cr_old

        This mirrors Abaqus's CETOL criterion, which bounds the maximum
        equivalent creep strain increment per increment to ensure accuracy of
        the explicit-in-time creep integration.

        Returns
        -------
        float
            Maximum equivalent creep strain increment (scalar, >= 0).
        """
        # int_vars layout: a1,a2,a3,a4,a5,a6,a7,a8,a9, a10(alpha_cr), a11, a12
        # Index 9 is alpha_cr (accumulated equivalent creep strain)
        alpha_cr_old = int_vars_old[9]   # shape (num_cells, num_quads)
        alpha_cr_new = int_vars_new[9]
        delta_alpha_cr = alpha_cr_new - alpha_cr_old
        return float(np.max(np.abs(delta_alpha_cr)))

    def update_shape_grads(self, sol):
        old_shape_grads = self.fe.shape_grads
        self.fe.shape_grads, self.fe.JxW = self.fe.get_shape_grads(sol)

    def compute_stress(self, sol, int_vars):
        self._ensure_compiled_maps()
        u_grads = (np.take(sol, self.fe.cells, axis=0)[:, None, :, :, None] *
                    self.fe.shape_grads[:, :, :, None, :])
        u_grads = np.sum(u_grads, axis=2)  # (num_cells, num_quads, vec, dim)
        sigma = self._compiled_cauchy_stress(u_grads, *int_vars)
        return sigma

    def compute_von_mises(self, s):
        # s: (num_cells, num_quads, 3, 3)  Cauchy stress
        # Flatten to (N, 3, 3), apply, reshape back so the cached vmap handles it.
        self._ensure_compiled_maps()
        flat = s.reshape(-1, 3, 3)
        vm_flat = self._compiled_von_mises(flat)
        return vm_flat.reshape(s.shape[:2])

    def compute_mag_logarithmic_strain(self, e):
        def logarithmic_strain(epsilon):
            w, v = np.linalg.eigh(2. * epsilon + np.eye(3))
            return np.abs(np.log(np.sqrt(np.abs(w[2]))))

        log_strain_fn = jax.vmap(logarithmic_strain)
        return log_strain_fn(e)

    def compute_strain(self, sol, int_vars):
        self._ensure_compiled_maps()
        u_grads = (np.take(sol, self.fe.cells, axis=0)[:, None, :, :, None] *
                    self.fe.shape_grads[:, :, :, None, :])
        u_grads = np.sum(u_grads, axis=2)  # (num_cells, num_quads, vec, dim)
        strain = self._compiled_lagrangian_strain(u_grads, *int_vars)
        return strain

    def compute_log_strain(self, sol, int_vars):
        self._ensure_compiled_maps()
        u_grads = (np.take(sol, self.fe.cells, axis=0)[:, None, :, :, None] *
                    self.fe.shape_grads[:, :, :, None, :])
        u_grads = np.sum(u_grads, axis=2)  # (num_cells, num_quads, vec, dim)
        strain = self._compiled_log_strain(u_grads, *int_vars)
        return strain

    def compute_nodal_values(self, int_point_values):
        s = int_point_values.shape
        # int_point_values_reshape = np.squeeze(int_point_values.reshape((1, -1, s[2], s[3])))
        int_point_values_reshape = int_point_values.reshape((s[0], s[1], -1))

        def shape_fun(input):
            xi_o, eta_o, zeta_o = input
            xi = 2 * xi_o - 1
            eta = 2 * eta_o - 1
            zeta = 2 * zeta_o - 1
            N1 = 1 / 8 * (1 - xi) * (1 - eta) * (1 - zeta)
            N2 = 1 / 8 * (1 + xi) * (1 - eta) * (1 - zeta)
            N3 = 1 / 8 * (1 + xi) * (1 + eta) * (1 - zeta)
            N4 = 1 / 8 * (1 - xi) * (1 + eta) * (1 - zeta)
            N5 = 1 / 8 * (1 - xi) * (1 - eta) * (1 + zeta)
            N6 = 1 / 8 * (1 + xi) * (1 - eta) * (1 + zeta)
            N7 = 1 / 8 * (1 + xi) * (1 + eta) * (1 + zeta)
            N8 = 1 / 8 * (1 - xi) * (1 + eta) * (1 + zeta)
            return np.array([N1, N2, N3, N4, N5, N6, N7, N8])

        extrapolation = np.array(jax.vmap(shape_fun)(np.squeeze(self.fe.quad_points_ref))).T

        def nodal(data):
            def nodal_vmap(data):
                return extrapolation @ data
            return jax.vmap(nodal_vmap, in_axes=1, out_axes=1)(data)
        nodal_values = jax.vmap(nodal, in_axes=0)(int_point_values_reshape)
        return nodal_values.reshape(s)

    def compute_pressure(self, sol, int_var):
        boundary_inds = self.traction_inds
        face_inds = self.fe.face_inds[boundary_inds[:, 1]]
        sigmas_integration_points = self.compute_stress(sol, int_var)
        sigmas_all = self.compute_nodal_values(sigmas_integration_points)
        sigmas = sigmas_all[boundary_inds[:, 0]]
        logging.info('Modified: Elem Stress %i' % boundary_inds[0][0])
        logging.info(sigmas_all.mean(axis=1)[boundary_inds[0][0]])
        ## Note: Face inds assumes nodal values are being used, not the elemental stress value
        sigmas = np.take_along_axis(sigmas_all[boundary_inds[:, 0]], face_inds[:, :, None, None], axis=1)
        logging.info('Modified: Face Stress1')
        logging.info(sigmas.mean(axis=1)[0])

        n_trial = self.fe.get_normal_vectors(boundary_inds, sol)

        sigma_n = np.squeeze(sigmas @ n_trial[:, None, :, None])
        pressure = np.squeeze(sigma_n @ n_trial[:, :, None])
        pressure = np.mean(pressure, axis=1)
        # p_avg = np.abs(np.mean(pressure[:, 0, 0]))
        return pressure, n_trial

    def compute_traction(self, location_fn, sol):
        """For post-processing only
        """
        boundary_inds = self.traction_inds

        if self.scale == 0:
            self.F_old_traction = np.repeat(
                np.repeat(np.eye(self.dim)[None, None, :, :], boundary_inds.shape[0], axis=0), 4, axis=1)
            self.Be_old_traction = np.array(self.F_old_traction)
            self.alpha_old_traction = np.zeros((boundary_inds.shape[0], 4))

        tensor_map = self.get_tensor_map()

        def traction_fn(self, u_grads, F_old, Be_old, alpha_old, boundary_inds):
            u_grads_reshape = u_grads.reshape(-1, self.vec[0], self.dim)
            F_old_reshape = F_old.reshape(-1, self.vec[0], self.dim)
            Be_old_reshape = Be_old.reshape(-1, self.vec[0], self.dim)
            alpha_old_reshape = alpha_old.reshape(-1)
            sigmas = jax.vmap(tensor_map)(u_grads_reshape, F_old_reshape, Be_old_reshape, alpha_old_reshape).reshape(
                u_grads.shape)

            n_trial = self.fe.get_normal_vectors(boundary_inds, sol)
            traction = (sigmas @ self.fe.init_norm[:, None, :, None])[:, :, :, 0]
            return traction, n_trial

        traction_elem, traction_integral_val, n_trial = self.surface_integral(boundary_inds, traction_fn, sol)
        return traction_elem, traction_integral_val, n_trial

    def surface_integral(self, boundary_inds, surface_fn, sol):
        self._ensure_compiled_maps()
        vmap_update_int_vars_map = self._compiled_update_int_vars
        # boundary_inds = self.get_boundary_conditions_inds([location_fn])[0]

        face_shape_grads_physical, nanson_scale = self.fe.get_face_shape_grads(boundary_inds, sol)
        # u_grads_face_i = sol[self.cells][boundary_inds[:, 0]][:, None, :, :, None] * face_shape_grads_physical[:, :, :, None, :]
        u_grads_face_i = (sol[self.fe.cells][boundary_inds[:, 0]][:, None, :, :, None] *
                            face_shape_grads_physical[:, :, :, None, :])
        u_grads_face_i = np.sum(u_grads_face_i, axis=2)
        updated_int_vars_traction = vmap_update_int_vars_map(u_grads_face_i, self.F_old_traction, self.Be_old_traction,
                                                            self.alpha_old_traction)  # (num_selected_faces, num_face_quads, vec, dim)
        traction, n_trial = surface_fn(self, u_grads_face_i, self.F_old_traction, self.Be_old_traction,
                                        self.alpha_old_traction,
                                        boundary_inds)  # (num_selected_faces, num_face_quads, vec)
        traction_elem = np.sum(traction * nanson_scale[:, :, None], axis=(1))
        self.F_old_traction, self.Be_old_traction, self.alpha_old_traction = updated_int_vars_traction
        int_val = np.sum(traction * nanson_scale[:, :, None], axis=(0, 1))
        return traction_elem, int_val, n_trial

    def set_params(self, params):
        int_vars, scale = params
        self.scale = scale
        self.internal_vars = int_vars
        # self.dirichlet_bc_info[-1][-3] = hemisphere_x(scale)
        self.fe.update_Dirichlet_boundary_conditions(self.fe.dirichlet_bc_info)

class PlasticCreepAxisy(Plasticity):
    def set_params(self, params):
        """
        set_params is invoked once per Newton iteration by the FEM driver, but
        within a single time step the inputs (scale, dt, int_vars, rho_ini) do
        not change -- only `sol` changes (and we use ref-config surface geometry).

        We therefore cache the heavy work (theta broadcast, scales array, surface
        normals, int_vars repacking) by a step-level key and short-circuit on
        repeated calls with the same key.
        """
        int_vars, scale, sol, rho_ini, dt = params

        # Build a tracer-safe cache key using Python object identity only.
        # `id()` works for both concrete and traced JAX arrays; `float()` /
        # `tobytes()` would raise TracerArrayConversionError under jax.vjp.
        #
        # The time-stepping core only swaps in fresh `int_vars`, `scale`, `dt`
        # at the END of an accepted step.  Within one Newton-iter cluster of
        # a single time step, all four objects are passed unchanged, so id()-
        # based identity is exactly what the cache needs.  Across optimizer
        # steps, all four ids change because the parent objects are rebuilt.
        step_key = (id(scale), id(dt), id(int_vars), id(rho_ini))

        if self._set_params_step_key != step_key:
            # Genuinely new time step / parameter set -- rebuild caches.
            self.dt = dt
            a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12 = int_vars

            self._cached_step_scales = scale * np.ones(
                (len(self.boundary_inds_list[0]), self.fes[0].num_face_quads)
            )

            ref_pt = np.array([1.5, -5.])
            norm_quad_surface = self._get_ref_normals(ref_pt, sol.shape)

            thetas = self._build_thetas(rho_ini)

            self._cached_step_int_vars = [a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, thetas, a12]
            self._cached_step_norm     = norm_quad_surface
            self._set_params_step_key  = step_key

        # Reference-config surface quad points are constant for a fixed mesh.
        physical_surface_quad_points = self._get_ref_surface_quad(sol.shape)

        self.internal_vars          = self._cached_step_int_vars
        self.internal_vars_surfaces = [[
            self._cached_step_scales,
            self._cached_step_norm,
            physical_surface_quad_points,
        ]]
    
    def set_init_params(self, params):
        int_vars, scale, sol, rho_ini, dt= params
        a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12 = int_vars
        self.rho_ini = rho_ini
        full_params = np.ones((self.fe.num_cells, len(rho_ini)))
        full_params = full_params.at[self.fe.flex_inds].set(rho_ini)
        self.thetas = np.repeat(full_params[:, None, :], self.fe.num_quads, axis=1)
        self.internal_vars = [a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, self.thetas, a12]